import openpyxl
import os
from common.constant import DATA_DIR
from common.config import conf

filename = conf.get('excel','file_name')
file_name = os.path.join(DATA_DIR,filename)

class Case:
    """存放用例"""
    pass


class CaseNew:
    def __init__(self, attrs):
        # 这个类用来存储用例
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """读取数据"""

    def __init__(self, file_name, sheet_name):
        """
        这个是用例初始化读取对象的
        :param file_name: 文件名字 --> str
        :param sheet_name: 表单名字 --> str
        """
        self.file_name = file_name
        self.sheet_name = sheet_name


    def open(self):
        # 打开工作薄
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sheet = self.wb[self.sheet_name]

    def close(self):
        # 关闭工作薄
        self.wb.close()

    def read_data_line(self):
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        for case in rows_data[1:]:
            # 获取一条测试用例
            data = []
            for cell in case:
                # if isinstance(cell.value, str):
                #     data.append(eval(cell.value))
                # else:
                    data.append(cell.value)
            # 将该条数据放入cases中
            case_data = dict(zip(titles, data))
            cases.append(case_data)
        self.close()
        return cases

    def read_data_line_obj(self):
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
            # 定义一个空列表用来存储所有的用例
            cases = []
        for case in rows_data[1:]:
            data = []
            for cell in case:
            # 获取一条测试用例
            #     if isinstance(cell.value, str):
            #         data.append(eval(cell.value))
            #     else:
                    data.append(cell.value)
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))
            case_obj = Case()
            for i in case_data:
                setattr(case_obj, i[0], i[1])
            cases.append(case_obj)
            # print(case_obj.id, case_obj.data)
        self.close()
        return cases

    def read_data_line_obj_new(self):
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
            # 定义一个空列表用来存储所有的用例
            cases = []
        for case in rows_data[1:]:
            data = []
            for cell in case:
                # 获取一条测试用例
                #     if isinstance(cell.value, str):
                #         data.append(eval(cell.value))
                #     else:
                data.append(cell.value)
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))
            case_obj = CaseNew(case_data)
            cases.append(case_obj)
            # print(case_obj.id, case_obj.data)
        self.close()
        return cases

    def read_data_list(self, list):
        self.open()
        # 获取最大行数
        max_r = self.sheet.max_row
        # 空列表存放所有用例
        cases = []
        # 空列表存放表头
        titles = []
        # 遍历所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row != 1:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                # 将该条数据和表头打包
                case = dict(zip(titles, case_data))
                cases.append(case)
            else:
                for column in list:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        self.close()
        return cases

    def read_data_list_obj(self, list1):
        self.open()
        # 获取最大行数
        max_r = self.sheet.max_row
        # 空列表存放所有用例
        cases = []
        # 空列表存放表头
        titles = []
        # 遍历所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row != 1:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                # 将该条数据和表头打包
                case = list(zip(titles, case_data))
                # 将一条用例存入一个对象中，每一列对应对象的一个属性
                case_obj = Case()
                for i in case:
                    setattr(case_obj,i[0],i[1])
                cases.append(case_obj)

            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        self.close()
        return cases

    def read_data_list_obj_new(self,list2):
        self.open()
        # 获取最大行数
        max_r = self.sheet.max_row
        # 空列表存放所有用例
        cases = []
        # 空列表存放表头
        titles = []
        # 遍历所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row != 1:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list2:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                # 将该条数据和表头打包
                case = list(zip(titles, case_data))
                case_obj = CaseNew(case)
                cases.append(case_obj)
            else:
                for column in list2:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        self.close()
        return cases

    def write_data(self, row, column, msg):
        self.open()
        self.sheet.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
        self.close()
        return print("写人完成")


if __name__ == '__main__':

    r = ReadExcel(file_name,'Sheet1')

    # res = r.read_data_list([1,2,3,4])
    # print(res)

    # res = r.read_data_list_obj([1, 2, 3, 4])
    # for i in res:
    #     print(i.case_id)
    #     print(i.excepted)

    # res1 = r.read_data_list_obj_new([1, 2, 3, 4])
    # for i in res1:
    #     print(i.case_id)
    #     print(i.excepted)

    # res = r.read_data_line()
    # for i in res:
    #     print(i['case_id'])
    #     print(i['testcase'])
    #     print(i['input'])

    # res = r.read_data_line_obj()
    # for i in res:
    #     print(i.case_id)
    #     print(i.excepted)

    # res = r.read_data_line_obj_new()
    # for i in res:
    #     print(i.case_id)
    #     print(i.excepted)

    # r.write_data(2,5,'1000')
    # r.__close__()